from __future__ import annotations

import os
import re
from copy import copy
import secrets
import sys
import mimetypes
from base64 import b64encode
from io import BytesIO
from datetime import UTC, datetime
from functools import wraps
from pathlib import Path
from typing import Any, Callable, TypeVar
from urllib import error as urllib_error
from urllib import request as urllib_request

BASE_DIR = Path(__file__).resolve().parent
LOCAL_VENDOR_DIR = BASE_DIR / "_vendor"
if LOCAL_VENDOR_DIR.exists():
    sys.path.insert(0, str(LOCAL_VENDOR_DIR))

from flask import Flask, g, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from psycopg.errors import ForeignKeyViolation
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash

from db import (
    category_key_exists,
    count_active_admin_users,
    count_admin_users,
    count_orders,
    count_products,
    count_store_users,
    count_units_in_stock,
    create_admin_session,
    create_admin_user,
    create_banner,
    create_category,
    create_order,
    delete_orders,
    create_product,
    create_products_batch,
    create_store_user,
    delete_admin_session,
    delete_admin_sessions_for_user,
    delete_admin_user,
    delete_banner,
    delete_category,
    delete_product,
    delete_store_user,
    ensure_database_ready,
    get_admin_user_by_email,
    get_admin_user_by_id,
    get_category_by_id,
    get_admin_user_by_session_token,
    get_banner_by_id,
    get_homepage_config,
    get_order_by_id,
    get_product_by_id,
    get_product_by_slug,
    get_store_user_by_email,
    get_store_user_by_id,
    list_admin_users,
    list_banners,
    list_categories,
    list_category_labels,
    list_orders,
    list_products,
    list_store_users,
    product_code_exists,
    product_sku_exists,
    product_slug_exists,
    save_homepage_config,
    update_admin_user,
    update_banner,
    update_category,
    update_order_status,
    update_product,
    update_product_inventory,
    update_store_user,
)

app = Flask(__name__)
CORS(app)
UPLOAD_DIR = BASE_DIR / "uploads"
ADMIN_FRONTEND_DIST = BASE_DIR.parent / "frontend" / "dist"
mimetypes.add_type("image/webp", ".webp")

F = TypeVar("F", bound=Callable[..., Any])
SERVICE_TOKEN = os.environ.get("LUMIERE_SERVICE_TOKEN", "lumiere-service-token")
PASSWORD_HASH_METHOD = "pbkdf2:sha256:600000"
SUPPORTED_LANGS = {"zh", "en"}
DEFAULT_LANG = "zh"
ORDER_STATUSES = {"pending_payment", "paid", "shipped", "completed", "cancelled"}
ORDER_STATUS_LABELS = {
    "pending_payment": "待付款",
    "paid": "已付款",
    "shipped": "已发货",
    "completed": "已完成",
    "cancelled": "已取消",
}


def parse_bool(value: str | None) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def matches_time_range(value: str, range_key: str) -> bool:
    if range_key == "all":
        return True
    try:
        target = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return False
    now = datetime.now(UTC)
    if range_key == "today":
        target_local = target.astimezone(UTC)
        return target_local.date() == now.date()
    if range_key == "yesterday":
        target_local = target.astimezone(UTC)
        return target_local.date() == (now.date()).fromordinal(now.date().toordinal() - 1)
    diff_days = (now - target.astimezone(UTC)).total_seconds() / 86400
    if range_key == "7d":
        return diff_days <= 7
    if range_key == "30d":
        return diff_days <= 30
    if range_key == "90d":
        return diff_days <= 90
    if range_key == "year":
        return target.astimezone(UTC).year == now.year
    return True


def filter_orders(
    orders: list[dict[str, Any]],
    *,
    time_range: str = "all",
    status: str = "all",
    category: str = "all",
    keyword: str = "",
) -> list[dict[str, Any]]:
    normalized_keyword = str(keyword or "").strip().lower()
    items: list[dict[str, Any]] = []
    for order in orders:
        status_match = status == "all" or order.get("status") == status
        time_match = matches_time_range(str(order.get("createdAt") or ""), time_range)
        category_match = category == "all" or any(
            str(item.get("categoryKey") or "") == category for item in (order.get("items") or [])
        )
        keyword_match = not normalized_keyword or normalized_keyword in " ".join(
            [
                str(order.get("orderNo") or "").lower(),
                str(order.get("userName") or "").lower(),
            ]
        )
        if status_match and time_match and category_match and keyword_match:
            items.append(order)
    return items


def fetch_image_bytes(url: str) -> bytes | None:
    value = str(url or "").strip()
    if not value:
        return None
    req = urllib_request.Request(
        value,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
        },
    )
    try:
        with urllib_request.urlopen(req, timeout=20) as resp:
            content_type = str(resp.headers.get("Content-Type", "")).lower()
            if not content_type.startswith("image/"):
                return None
            return resp.read()
    except Exception:
        return None


def build_excel_image(image_bytes: bytes, *, width: int = 54, height: int = 70) -> OpenpyxlImage | None:
    try:
        with PILImage.open(BytesIO(image_bytes)) as img:
            converted = img.convert("RGBA")
            output = BytesIO()
            converted.save(output, format="PNG")
            output.seek(0)
            excel_image = OpenpyxlImage(output)
            excel_image.width = width
            excel_image.height = height
            return excel_image
    except Exception:
        return None


def autosize_columns(worksheet: Any) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 12), 42)


def set_fixed_column_widths(worksheet: Any, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[str(column).upper()].width = float(width)


def estimate_text_row_height(
    value: Any,
    *,
    line_width: int = 28,
    min_height: float = 24.0,
    line_height: float = 18.0,
    max_height: float = 120.0,
) -> float:
    text = str(value or "").strip()
    if not text:
        return min_height

    visual_lines = 0
    for raw_line in text.splitlines() or [""]:
        line = raw_line.strip()
        if not line:
            visual_lines += 1
            continue
        visual_lines += max(1, (len(line) + line_width - 1) // line_width)

    return min(max_height, max(min_height, visual_lines * line_height))


def is_image_attachment(url: str) -> bool:
    value = str(url or "").strip().lower()
    return bool(re.search(r"\.(?:jpg|jpeg|png|webp)(?:$|[?#])", value))


def split_order_attachments(order: dict[str, Any]) -> tuple[list[str], list[str]]:
    raw = order.get("labelImageUrls") or ([order.get("labelPdfUrl")] if order.get("labelPdfUrl") else [])
    image_urls: list[str] = []
    file_urls: list[str] = []
    for item in raw:
        url = str(item or "").strip()
        if not url:
            continue
        if is_image_attachment(url):
            image_urls.append(url)
        else:
            file_urls.append(url)
    return image_urls[:5], file_urls[:5]


def build_orders_export(orders: list[dict[str, Any]], *, include_images: bool = True) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders"
    headers = [
        "订单号",
        "业务员",
        "客户邮箱",
        "联系人",
        "联系电话",
        "收货地址",
        "商品图片",
        "SKU",
        "尺码",
        "数量",
        "备注",
        "备注附件",
        "附件图片1",
        "附件图片2",
        "附件图片3",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center")

    current_row = 2
    for order in orders:
        rows = order.get("items") or [{}]
        attachment_images, attachment_files = split_order_attachments(order)
        for item in rows:
            worksheet.append(
                [
                    order.get("orderNo") or "",
                    order.get("userName") or "",
                    order.get("userEmail") or "",
                    order.get("contactName") or "",
                    order.get("phone") or "",
                    order.get("shippingAddress") or "",
                    "",
                    item.get("sku") or "",
                    item.get("sizeCode") or "",
                    item.get("quantity") or 0,
                    order.get("note") or "",
                    ", ".join(attachment_files),
                    "",
                    "",
                    "",
                ]
            )
            worksheet.row_dimensions[current_row].height = 118
            for cell in worksheet[current_row]:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            if include_images:
                image_bytes = fetch_image_bytes(str(item.get("image") or ""))
                if image_bytes:
                    try:
                        excel_image = build_excel_image(image_bytes, width=86, height=112)
                        if excel_image:
                            worksheet.add_image(excel_image, f"G{current_row}")
                    except Exception:
                        pass
                for attachment_index, attachment_url in enumerate(attachment_images[:3], start=13):
                    try:
                        attachment_bytes = fetch_image_bytes(attachment_url)
                        if not attachment_bytes:
                            continue
                        attachment_image = build_excel_image(attachment_bytes, width=86, height=86)
                        if attachment_image:
                            worksheet.add_image(attachment_image, f"{get_column_letter(attachment_index)}{current_row}")
                    except Exception:
                        continue

            current_row += 1

    worksheet.freeze_panes = "A2"
    autosize_columns(worksheet)
    worksheet.column_dimensions["G"].width = 16
    worksheet.column_dimensions["L"].width = max(float(worksheet.column_dimensions["L"].width or 0), 28)
    worksheet.column_dimensions["M"].width = 16
    worksheet.column_dimensions["N"].width = 16
    worksheet.column_dimensions["O"].width = 16

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

PROFORMA_TEMPLATE_PATH = Path(
    os.environ.get("PROFORMA_TEMPLATE_PATH", r"E:\PROFORMA INVIOCE-GINGTTO to dear  -20260211.xlsx")
).expanduser()


def copy_row_style(worksheet: Any, source_row: int, target_row: int, max_col: int = 6) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column in range(1, max_col + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def normalize_order_shipping_fee(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def build_invoice_party_name(order: dict[str, Any]) -> str:
    return (
        str(order.get("companyName") or "").strip()
        or str(order.get("contactName") or "").strip()
        or str(order.get("userName") or "").strip()
        or "--"
    )


def build_invoice_email(order: dict[str, Any]) -> str:
    return str(order.get("contactValue") or "").strip() or str(order.get("userEmail") or "").strip()


def build_invoice_address(order: dict[str, Any]) -> str:
    explicit = str(order.get("shippingAddress") or "").strip()
    if explicit:
        return explicit
    parts = [
        str(order.get("address") or "").strip(),
        str(order.get("apartment") or "").strip(),
        str(order.get("city") or "").strip(),
        str(order.get("state") or "").strip(),
        str(order.get("zip") or "").strip(),
        str(order.get("country") or "").strip(),
    ]
    return ", ".join(part for part in parts if part)


def build_invoice_item_description(item: dict[str, Any]) -> str:
    sku = str(item.get("sku") or "").strip()
    return sku or "--"


def safe_sheet_title(value: Any, fallback: str = "Sheet") -> str:
    text = str(value or "").strip() or fallback
    text = text.translate({ord(ch): "_" for ch in "[]:*?/\\"})
    return text[:31]


def build_orders_sheet_export(orders: list[dict[str, Any]], *, include_images: bool = True) -> BytesIO:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    if not orders:
        empty_sheet = workbook.create_sheet(title="Orders")
        empty_sheet["A1"] = "No orders"

    for index, order in enumerate(orders, start=1):
        order_no = str(order.get("orderNo") or f"order_{index}").strip() or f"order_{index}"
        worksheet = workbook.create_sheet(title=safe_sheet_title(order_no, f"order_{index}"))
        attachment_images, attachment_files = split_order_attachments(order)
        owner_name = str(order.get("userName") or "--").strip() or "--"

        worksheet.sheet_view.showGridLines = False
        set_fixed_column_widths(
            worksheet,
            {
                "A": 24,
                "B": 19,
                "C": 4,
                "D": 14,
                "E": 4,
                "F": 14,
            },
        )

        worksheet["A1"] = owner_name
        worksheet["A1"].font = Font(size=20, bold=True)
        worksheet["A1"].alignment = Alignment(vertical="center")
        worksheet["F1"] = "快递"
        worksheet["F1"].font = Font(size=18, bold=True)
        worksheet["F1"].alignment = Alignment(horizontal="right", vertical="center")
        worksheet.row_dimensions[1].height = 28

        worksheet["A3"] = "客户："
        worksheet["B3"] = str(order.get("contactName") or "--").strip() or "--"
        worksheet["A4"] = "电话："
        worksheet["B4"] = str(order.get("phone") or "--").strip() or "--"
        worksheet["A5"] = "地址："
        worksheet["B5"] = build_invoice_address(order) or "--"
        worksheet["A6"] = "订单号："
        worksheet["B6"] = order_no

        for label_cell in ("A3", "A4", "A5", "A6"):
            worksheet[label_cell].font = Font(bold=True)
            worksheet[label_cell].alignment = Alignment(horizontal="right", vertical="center")

        for value_cell in ("B3", "B4", "B6"):
            worksheet[value_cell].alignment = Alignment(vertical="center")

        worksheet.merge_cells("B5:F5")
        worksheet["B5"].alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.row_dimensions[3].height = 22
        worksheet.row_dimensions[4].height = 22
        worksheet.row_dimensions[5].height = estimate_text_row_height(worksheet["B5"].value, line_width=42, min_height=40, line_height=18, max_height=88)
        worksheet.row_dimensions[6].height = 22

        table_row = 8
        header_fill = PatternFill(fill_type="solid", fgColor="FFF36A")
        headers = [("A", "型号规格"), ("B", "图片"), ("D", "尺码"), ("F", "数量")]
        for column, title in headers:
            cell = worksheet[f"{column}{table_row}"]
            cell.value = title
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[table_row].height = 24

        items = [item for item in (order.get("items") or []) if isinstance(item, dict)] or [{}]
        current_row = table_row + 1
        total_quantity = 0
        for item in items:
            quantity = int(item.get("quantity") or 0)
            total_quantity += quantity
            worksheet[f"A{current_row}"] = str(item.get("sku") or item.get("productName") or "--").strip() or "--"
            worksheet[f"D{current_row}"] = str(item.get("sizeCode") or "--").strip() or "--"
            worksheet[f"F{current_row}"] = quantity
            worksheet[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            worksheet[f"D{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
            worksheet[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
            worksheet.row_dimensions[current_row].height = 86
            if include_images:
                image_bytes = fetch_image_bytes(str(item.get("image") or ""))
                if image_bytes:
                    excel_image = build_excel_image(image_bytes, width=104, height=82)
                    if excel_image:
                        worksheet.add_image(excel_image, f"B{current_row}")
            current_row += 1

        summary_row = current_row
        worksheet[f"D{summary_row}"] = "总件数："
        worksheet[f"F{summary_row}"] = total_quantity
        worksheet[f"D{summary_row}"].font = Font(bold=True)
        worksheet[f"F{summary_row}"].font = Font(bold=True)
        worksheet[f"D{summary_row}"].alignment = Alignment(horizontal="right", vertical="center")
        worksheet[f"F{summary_row}"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[summary_row].height = 24

        footer_row = summary_row + 2
        worksheet[f"A{footer_row}"] = "备注："
        worksheet[f"A{footer_row}"].font = Font(bold=True)
        worksheet.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=6)
        worksheet[f"B{footer_row}"] = str(order.get("note") or "").strip() or "--"
        worksheet[f"B{footer_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.row_dimensions[footer_row].height = estimate_text_row_height(worksheet[f"B{footer_row}"].value, line_width=58, min_height=34, line_height=18, max_height=96)

        attachment_text_row = footer_row + 1
        if attachment_files:
            worksheet[f"A{attachment_text_row}"] = "附件："
            worksheet[f"A{attachment_text_row}"].font = Font(bold=True)
            worksheet.merge_cells(start_row=attachment_text_row, start_column=2, end_row=attachment_text_row, end_column=6)
            worksheet[f"B{attachment_text_row}"] = "\n".join(attachment_files)
            worksheet[f"B{attachment_text_row}"].alignment = Alignment(wrap_text=True, vertical="top")
            worksheet.row_dimensions[attachment_text_row].height = estimate_text_row_height(worksheet[f"B{attachment_text_row}"].value, line_width=58, min_height=28, line_height=18, max_height=90)

        if attachment_images:
            image_label_row = attachment_text_row + 2
            worksheet[f"A{image_label_row}"] = "附件图片："
            worksheet[f"A{image_label_row}"].font = Font(bold=True)
            worksheet.row_dimensions[image_label_row].height = 22

            attachment_anchor_columns = ["B", "D", "F"]
            attachment_anchor_rows = [image_label_row + 1, image_label_row + 6]
            for anchor_row in attachment_anchor_rows:
                worksheet.row_dimensions[anchor_row].height = 82

            for image_index, image_url in enumerate(attachment_images[:6]):
                image_bytes = fetch_image_bytes(image_url)
                if not image_bytes:
                    continue
                attachment_image = build_excel_image(image_bytes, width=84, height=84)
                if not attachment_image:
                    continue
                anchor_col = attachment_anchor_columns[image_index % len(attachment_anchor_columns)]
                anchor_row = attachment_anchor_rows[image_index // len(attachment_anchor_columns)]
                worksheet.add_image(attachment_image, f"{anchor_col}{anchor_row}")

        worksheet.freeze_panes = "A8"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def reset_invoice_summary_merges(worksheet: Any, shipping_row: int, total_row: int, payment_row: int) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_col == 3 and merged_range.max_col == 5 and merged_range.min_row >= 16:
            worksheet.unmerge_cells(str(merged_range))
    for row in (shipping_row, total_row, payment_row):
        worksheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)


def build_order_invoice_export(order: dict[str, Any]) -> BytesIO:
    if not PROFORMA_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"未找到形式发票导出模板: {PROFORMA_TEMPLATE_PATH}")

    shipping_fee = normalize_order_shipping_fee(order.get("shippingFee"))
    if shipping_fee <= 0:
        raise ValueError("请先填写运费后再导出 PI")

    workbook = load_workbook(PROFORMA_TEMPLATE_PATH)
    worksheet = workbook["PI"] if "PI" in workbook.sheetnames else workbook.active
    worksheet.sheet_view.showGridLines = False

    items = [item for item in (order.get("items") or []) if isinstance(item, dict)]
    if not items:
        items = [{}]

    item_start_row = 13
    template_item_rows = 3
    template_last_item_row = item_start_row + template_item_rows - 1
    required_item_rows = max(len(items), template_item_rows)
    extra_rows = required_item_rows - template_item_rows

    if extra_rows > 0:
        insert_at = template_last_item_row + 1
        worksheet.insert_rows(insert_at, extra_rows)
        for offset in range(extra_rows):
            copy_row_style(worksheet, template_last_item_row, insert_at + offset)

    shipping_row = item_start_row + required_item_rows
    total_row = shipping_row + 1
    payment_row = shipping_row + 2

    reset_invoice_summary_merges(worksheet, shipping_row, total_row, payment_row)

    worksheet["B5"] = str(order.get("orderNo") or "")
    worksheet["E5"] = datetime.now().strftime("%Y.%m.%d")
    worksheet["B6"] = build_invoice_party_name(order)
    worksheet["E6"] = build_invoice_email(order)
    worksheet["E7"] = str(order.get("phone") or "").strip()
    worksheet["B9"] = build_invoice_address(order)
    worksheet["B9"].alignment = copy(worksheet["B6"].alignment)

    set_fixed_column_widths(
        worksheet,
        {
            "B": max(float(worksheet.column_dimensions["B"].width or 0), 19),
            "C": max(float(worksheet.column_dimensions["C"].width or 0), 10),
            "D": max(float(worksheet.column_dimensions["D"].width or 0), 10),
            "E": max(float(worksheet.column_dimensions["E"].width or 0), 12),
            "F": max(float(worksheet.column_dimensions["F"].width or 0), 12),
        },
    )
    worksheet.row_dimensions[9].height = estimate_text_row_height(worksheet["B9"].value, line_width=34, min_height=34, line_height=18, max_height=72)

    for row in range(item_start_row, shipping_row):
        worksheet.row_dimensions[row].height = max(worksheet.row_dimensions[row].height or 0, 92)
        for column in range(1, 7):
            worksheet.cell(row, column).value = None

    for index, item in enumerate(items):
        row = item_start_row + index
        quantity = int(item.get("quantity") or 0)
        unit_price = float(item.get("unitPrice") or 0)
        total_price = float(item.get("totalPrice") or 0) or quantity * unit_price
        worksheet[f"A{row}"] = build_invoice_item_description(item)
        worksheet[f"C{row}"] = str(item.get("sizeCode") or "").strip() or "--"
        worksheet[f"D{row}"] = quantity
        worksheet[f"E{row}"] = unit_price
        worksheet[f"F{row}"] = total_price
        worksheet[f"A{row}"].alignment = copy(worksheet[f"A{template_last_item_row}"].alignment)
        worksheet[f"C{row}"].alignment = copy(worksheet[f"C{template_last_item_row}"].alignment)
        worksheet[f"D{row}"].alignment = copy(worksheet[f"D{template_last_item_row}"].alignment)
        worksheet[f"E{row}"].alignment = copy(worksheet[f"E{template_last_item_row}"].alignment)
        worksheet[f"F{row}"].alignment = copy(worksheet[f"F{template_last_item_row}"].alignment)

        image_bytes = fetch_image_bytes(str(item.get("image") or ""))
        if image_bytes:
            excel_image = build_excel_image(image_bytes, width=82, height=82)
            if excel_image:
                worksheet.add_image(excel_image, f"B{row}")

    worksheet[f"B{shipping_row}"] = "Shipping Cost"
    worksheet[f"F{shipping_row}"] = shipping_fee
    worksheet[f"B{total_row}"] = "Total"
    worksheet[f"F{total_row}"] = f"=SUM(F{item_start_row}:F{shipping_row})"
    worksheet[f"B{payment_row}"] = "Payment"
    worksheet[f"F{payment_row}"] = f"=F{total_row}"

    attachment_images, attachment_files = split_order_attachments(order)
    remarks = []
    if str(order.get("note") or "").strip():
        remarks.append(f"Note: {str(order.get('note') or '').strip()}")
    if attachment_files:
        remarks.append("Attachments: " + ", ".join(attachment_files))
    if remarks:
        worksheet["A21"] = "\n".join(remarks)
        worksheet["A21"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.row_dimensions[21].height = estimate_text_row_height(worksheet["A21"].value, line_width=70, min_height=28, line_height=18, max_height=88)

    if attachment_images:
        worksheet["A22"] = "Attachment Images"
        worksheet["A22"].font = Font(bold=True)
        worksheet.row_dimensions[22].height = 22

    attachment_anchor_rows = [23, 29]
    attachment_anchor_columns = ["A", "C", "E"]
    for anchor_row in attachment_anchor_rows:
        worksheet.row_dimensions[anchor_row].height = 82

    for image_index, attachment_url in enumerate(attachment_images[:6]):
        try:
            attachment_bytes = fetch_image_bytes(attachment_url)
            if not attachment_bytes:
                continue
            attachment_image = build_excel_image(attachment_bytes, width=92, height=92)
            if not attachment_image:
                continue
            row = attachment_anchor_rows[image_index // len(attachment_anchor_columns)]
            column = attachment_anchor_columns[image_index % len(attachment_anchor_columns)]
            worksheet.add_image(attachment_image, f"{column}{row}")
        except Exception:
            continue

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


HOME_SECTION_KEYS = ("bestSeller", "newArrival", "specialPrice")
ADMIN_USER_ROLES = {"admin", "sales", "warehouse", "customer"}
CLOUDINARY_CLOUD_NAME = os.environ.get("CLOUDINARY_CLOUD_NAME", "").strip()
CLOUDINARY_API_KEY = os.environ.get("CLOUDINARY_API_KEY", "").strip()
CLOUDINARY_API_SECRET = os.environ.get("CLOUDINARY_API_SECRET", "").strip()
CLOUDINARY_FOLDER = os.environ.get("CLOUDINARY_UPLOAD_FOLDER", "gingtto").strip() or "gingtto"


def utc_now() -> str:
    return datetime.now(UTC).isoformat()


def pick_language() -> str:
    lang = request.args.get("lang", request.headers.get("X-Lang", DEFAULT_LANG))
    return lang if lang in SUPPORTED_LANGS else DEFAULT_LANG


def localize(value: Any, lang: str) -> Any:
    if isinstance(value, dict) and SUPPORTED_LANGS & set(value.keys()):
        return value.get(lang) or value.get(DEFAULT_LANG) or next(iter(value.values()))
    if isinstance(value, list):
        return [localize(item, lang) for item in value]
    if isinstance(value, dict):
        return {key: localize(item, lang) for key, item in value.items()}
    return value


def serialize_product(product: dict[str, Any], lang: str) -> dict[str, Any]:
    category_label = product.get("categoryLabel") or product.get("categoryKey", "")
    name = localize(product["name"], lang)
    summary = localize(product["summary"], lang)
    return {
        "id": product["id"],
        "slug": product["slug"],
        "sku": product["sku"],
        "productCode": product.get("productCode", ""),
        "colorGroup": product.get("colorGroup", ""),
        "colorName": product.get("colorName", ""),
        "colorHex": product.get("colorHex", ""),
        "categoryKey": product["categoryKey"],
        "categoryLabel": category_label,
        "price": product["price"],
        "formattedPrice": f"${product['price']}",
        "stock": product["stock"],
        "featured": bool(product.get("featured")),
        "origin": product.get("origin", ""),
        "sizes": product.get("sizes", []),
        "sizePrices": product.get("sizePrices", []),
        "image": product["image"],
        "gallery": product.get("gallery", []),
        "sizeChartImage": product.get("sizeChartImage", ""),
        "descriptionImage": product.get("descriptionImage", ""),
        "colorOptions": product.get("colorOptions", []),
        "name": name,
        "summary": summary,
        "description": localize(product["description"], lang),
        "searchText": " ".join([name, summary, category_label, product["sku"], product.get("productCode", ""), product.get("colorName", "")]).lower(),
    }

def serialize_banner(banner: dict[str, Any], lang: str) -> dict[str, Any]:
    return {
        "id": banner["id"],
        "image": banner["image"],
        "ctaPath": banner.get("ctaPath", "/shop"),
        "title": localize(banner["title"], lang),
        "subtitle": localize(banner["subtitle"], lang),
        "ctaLabel": localize(banner["ctaLabel"], lang),
    }


def sanitize_admin_user(user: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": user["id"],
        "name": user["name"],
        "email": user["email"],
        "role": user.get("role", "admin"),
        "status": user["status"],
        "createdAt": user["createdAt"],
    }


def sanitize_store_user(user: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": user["id"],
        "name": user["name"],
        "companyName": user.get("companyName", ""),
        "email": user["email"],
        "status": user["status"],
        "createdAt": user["createdAt"],
    }


def extract_token() -> str:
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        return auth_header.replace("Bearer ", "", 1).strip()
    return ""


def require_auth(func: F) -> F:
    @wraps(func)
    def wrapper(*args: Any, **kwargs: Any) -> Any:
        token = extract_token()
        user = get_admin_user_by_session_token(token)
        if not user:
            return jsonify({"message": "Unauthorized"}), 401
        g.current_user = user
        return func(*args, **kwargs)

    return wrapper  # type: ignore[return-value]


def require_roles(*roles: str) -> Callable[[F], F]:
    allowed_roles = {role.strip().lower() for role in roles if role.strip()}

    def decorator(func: F) -> F:
        @wraps(func)
        def wrapper(*args: Any, **kwargs: Any) -> Any:
            current_role = str(g.current_user.get("role", "admin")).strip().lower()
            if current_role not in allowed_roles:
                return jsonify({"message": "Forbidden"}), 403
            return func(*args, **kwargs)

        return wrapper  # type: ignore[return-value]

    return decorator


def require_service_auth(func: F) -> F:
    @wraps(func)
    def wrapper(*args: Any, **kwargs: Any) -> Any:
        if request.headers.get("X-Service-Token", "") != SERVICE_TOKEN:
            return jsonify({"message": "Unauthorized service"}), 401
        return func(*args, **kwargs)

    return wrapper  # type: ignore[return-value]


def normalize_admin_role(raw_value: Any, default: str = "admin") -> str:
    role = str(raw_value or default).strip().lower() or default
    if role not in ADMIN_USER_ROLES:
        raise ValueError("Invalid role")
    return role


def validate_product_payload(payload: dict[str, Any]) -> str | None:
    if not str(payload.get("categoryKey", "")).strip():
        return "Missing field: categoryKey"
    if not str(payload.get("title", "")).strip():
        return "Missing field: title"

    if not str(payload.get("sizeChartImage", "")).strip():
        return "Missing field: sizeChartImage"
    if not str(payload.get("descriptionImage", "")).strip():
        return "Missing field: descriptionImage"

    sizes = [str(item).strip() for item in payload.get("sizes", []) if str(item).strip()]
    if not sizes:
        return "Missing field: sizes"


    variants = payload.get("variants")
    if variants is None:
        for field in ["productCode", "slug", "sku", "image"]:
            if not str(payload.get(field, "")).strip():
                return f"Missing field: {field}"
        size_prices = payload.get("sizePrices") or []
        if len(size_prices) != len(sizes):
            return "Missing field: sizePrices"
        for index, item in enumerate(size_prices, start=1):
            if not str(item.get("sizeCode", "")).strip() or item.get("price") in (None, "") or item.get("stock") in (None, ""):
                return f"Invalid sizePrices row at index {index}"
        return None

    if not isinstance(variants, list) or not variants:
        return "Missing field: variants"
    if not str(payload.get("familyCode", "")).strip() and not str(payload.get("colorGroup", "")).strip():
        return "Missing field: familyCode"

    for index, variant in enumerate(variants, start=1):
        if not isinstance(variant, dict):
            return f"Invalid variant at index {index}"
        if not str(variant.get("colorName", "")).strip():
            return f"Missing field: variants[{index}].colorName"
        if not str(variant.get("colorHex", "")).strip():
            return f"Missing field: variants[{index}].colorHex"
        image_urls = variant.get("imageUrls") or variant.get("gallery") or []
        if not isinstance(image_urls, list) or not image_urls or len(image_urls) > 10:
            return f"Invalid field: variants[{index}].imageUrls"
        size_prices = variant.get("sizePrices") or []
        if not isinstance(size_prices, list) or len(size_prices) != len(sizes):
            return f"Missing field: variants[{index}].sizePrices"
        for size_index, size_price in enumerate(size_prices, start=1):
            if (
                not str(size_price.get("sizeCode", "")).strip()
                or size_price.get("price") in (None, "")
                or size_price.get("stock") in (None, "")
            ):
                return f"Invalid sizePrices row at index {index}-{size_index}"
    return None


def apply_default_category(payload: dict[str, Any]) -> None:
    if str(payload.get("categoryKey", "")).strip():
        return
    categories = list_categories()
    if categories:
        payload["categoryKey"] = categories[0]["key"]

def validate_banner_payload(payload: dict[str, Any]) -> str | None:
    if not str(payload.get("image", "")).strip():
        return "Missing field: image"
    for field in ["title", "subtitle", "ctaLabel"]:
        bundle = payload.get(field, {})
        if not isinstance(bundle, dict):
            return f"Invalid field: {field}"
        for lang in ["zh", "en"]:
            if not str(bundle.get(lang, "")).strip():
                return f"Missing field: {field}.{lang}"
    return None


def validate_category_payload(payload: dict[str, Any]) -> str | None:
    key = str(payload.get("key") or payload.get("categoryKey") or "").strip().lower()
    if not key:
        return "Missing field: key"
    labels = payload.get("labels") or payload.get("name")
    if not isinstance(labels, dict):
        return "Invalid field: labels"
    for lang in ("zh", "en"):
        if not str(labels.get(lang, "")).strip():
            return f"Missing field: labels.{lang}"
    try:
        int(payload.get("sortOrder") or 0)
    except (TypeError, ValueError):
        return "Invalid field: sortOrder"
    return None


def validate_homepage_config_payload(payload: dict[str, Any]) -> str | None:
    hero_banners = payload.get("heroBanners")
    section_product_ids = payload.get("sectionProductIds")
    collection_product_ids = payload.get("collectionProductIds")
    display_category_keys = payload.get("displayCategoryKeys")

    if not isinstance(hero_banners, dict):
        return "Invalid field: heroBanners"
    if not isinstance(section_product_ids, dict):
        return "Invalid field: sectionProductIds"
    if not isinstance(collection_product_ids, dict):
        return "Invalid field: collectionProductIds"
    if not isinstance(display_category_keys, list):
        return "Invalid field: displayCategoryKeys"
    if len(display_category_keys) > 5:
        return "Display categories cannot exceed 5"

    valid_product_ids = {int(item["id"]) for item in list_products()}
    valid_category_keys = {str(item["key"]) for item in list_categories()}

    for key in HOME_SECTION_KEYS:
        hero_image = str(hero_banners.get(key, "") or "").strip()
        if not hero_image:
            return f"Missing field: heroBanners.{key}"

        product_ids = section_product_ids.get(key)
        if not isinstance(product_ids, list):
            return f"Invalid field: sectionProductIds.{key}"
        if len(product_ids) > 5:
            return f"{key} home products cannot exceed 5"
        for product_id in product_ids:
            try:
                normalized_product_id = int(product_id)
            except (TypeError, ValueError):
                return f"Invalid product id for {key}"
            if normalized_product_id not in valid_product_ids:
                return f"Product not found for {key}"

        collection_ids = collection_product_ids.get(key)
        if not isinstance(collection_ids, list):
            return f"Invalid field: collectionProductIds.{key}"
        for product_id in collection_ids:
            try:
                normalized_product_id = int(product_id)
            except (TypeError, ValueError):
                return f"Invalid collection product id for {key}"
            if normalized_product_id not in valid_product_ids:
                return f"Collection product not found for {key}"

    for key in display_category_keys:
        normalized_key = str(key or "").strip().lower()
        if normalized_key not in valid_category_keys:
            return f"Category not found: {normalized_key}"

    return None


def build_homepage_payload(lang: str) -> dict[str, Any]:
    config = get_homepage_config()
    products = list_products()
    product_map = {int(item["id"]): item for item in products}
    category_map = {item["key"]: item["label"] for item in list_category_labels(lang)}

    selected_banners = []
    for key in HOME_SECTION_KEYS:
        image_url = str(config["heroBanners"].get(key, "") or "").strip()
        if not image_url:
            continue
        selected_banners.append(
            {
                "id": 0,
                "slotKey": key,
                "image": image_url,
                "ctaPath": "/shop",
                "title": "",
                "subtitle": "",
                "ctaLabel": "",
            }
        )

    sections = {}
    for key in HOME_SECTION_KEYS:
        section_items = []
        for product_id in config["sectionProductIds"].get(key, []):
            product = product_map.get(int(product_id))
            if product:
                section_items.append(serialize_product(product, lang))
        sections[key] = section_items[:5]

    categories = [
        {"key": key, "label": category_map[key]}
        for key in config["displayCategoryKeys"]
        if key in category_map
    ]

    stats = [
        {"label": {"zh": "在线 SKU", "en": "Live SKUs"}[lang], "value": str(count_products())},
        {"label": {"zh": "现货库存", "en": "Units in stock"}[lang], "value": str(count_units_in_stock())},
    ]

    return {
        "banners": selected_banners,
        "sections": sections,
        "categories": categories,
        "stats": stats,
        "featured": sections.get("bestSeller", []),
    }


def build_upload_url(filename: str) -> str:
    return f"{request.host_url.rstrip('/')}/uploads/{filename}"


def cloudinary_enabled() -> bool:
    return bool(CLOUDINARY_CLOUD_NAME and CLOUDINARY_API_KEY and CLOUDINARY_API_SECRET)


def build_multipart_form_data(fields: dict[str, str], file_name: str, file_bytes: bytes, content_type: str) -> tuple[bytes, str]:
    boundary = f"----CloudinaryBoundary{secrets.token_hex(12)}"
    body = bytearray()

    for key, value in fields.items():
        body.extend(f"--{boundary}\r\n".encode("utf-8"))
        body.extend(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"))
        body.extend(str(value).encode("utf-8"))
        body.extend(b"\r\n")

    body.extend(f"--{boundary}\r\n".encode("utf-8"))
    body.extend(
        f'Content-Disposition: form-data; name="file"; filename="{file_name}"\r\n'.encode("utf-8")
    )
    body.extend(f"Content-Type: {content_type}\r\n\r\n".encode("utf-8"))
    body.extend(file_bytes)
    body.extend(b"\r\n")
    body.extend(f"--{boundary}--\r\n".encode("utf-8"))

    return bytes(body), boundary


def upload_file_to_cloudinary(file_storage: Any) -> str:
    if not cloudinary_enabled():
        raise RuntimeError("Cloudinary is not configured")

    file_name = secure_filename(file_storage.filename or "") or f"upload-{secrets.token_hex(4)}.jpg"
    content_type = file_storage.mimetype or "application/octet-stream"
    file_bytes = file_storage.read()
    file_storage.stream.seek(0)
    if not file_bytes:
        raise ValueError("Empty file")

    endpoint = f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD_NAME}/image/upload"
    fields = {
        "folder": CLOUDINARY_FOLDER,
        "use_filename": "true",
        "unique_filename": "true",
        "overwrite": "false",
    }
    payload, boundary = build_multipart_form_data(fields, file_name, file_bytes, content_type)
    auth_token = b64encode(f"{CLOUDINARY_API_KEY}:{CLOUDINARY_API_SECRET}".encode("utf-8")).decode("ascii")
    request_obj = urllib_request.Request(
        endpoint,
        data=payload,
        headers={
            "Authorization": f"Basic {auth_token}",
            "Content-Type": f"multipart/form-data; boundary={boundary}",
        },
        method="POST",
    )

    try:
        with urllib_request.urlopen(request_obj, timeout=60) as response:
            data = response.read().decode("utf-8")
    except urllib_error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"Cloudinary upload failed: {detail or exc.reason}") from exc
    except urllib_error.URLError as exc:
        raise RuntimeError(f"Cloudinary upload failed: {exc.reason}") from exc

    try:
        import json

        parsed = json.loads(data)
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("Cloudinary upload failed: invalid response") from exc

    secure_url = str(parsed.get("secure_url", "")).strip()
    if not secure_url:
        raise RuntimeError("Cloudinary upload failed: missing secure_url")
    return secure_url


def save_file_locally(file_storage: Any) -> str:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    stem = secure_filename(Path(file_storage.filename or "").stem) or "upload"
    suffix = Path(file_storage.filename or "").suffix.lower() or ".jpg"
    unique_name = f"{stem}-{secrets.token_hex(6)}{suffix}"
    file_storage.save(UPLOAD_DIR / unique_name)
    return build_upload_url(unique_name)


def admin_frontend_ready() -> bool:
    return (ADMIN_FRONTEND_DIST / "index.html").exists()


@app.get("/uploads/<path:filename>")
def serve_upload(filename: str) -> Any:
    return send_from_directory(UPLOAD_DIR, filename)


@app.post("/api/admin/uploads")
@require_auth
@require_roles("admin", "sales")
def upload_files() -> Any:
    files = request.files.getlist("files")
    if not files:
        return jsonify({"message": "Missing files"}), 400
    urls: list[str] = []
    for file in files:
        if not file.filename:
            continue
        if cloudinary_enabled():
            urls.append(upload_file_to_cloudinary(file))
        else:
            urls.append(save_file_locally(file))
    return jsonify({"urls": urls})


@app.get("/api/health")
def health() -> Any:
    return jsonify({"status": "ok", "timestamp": utc_now()})


@app.post("/api/auth/admin/login")
def admin_login() -> Any:
    payload = request.get_json(silent=True) or {}
    email = str(payload.get("email", "")).strip().lower()
    password = str(payload.get("password", "")).strip()
    user = get_admin_user_by_email(email, include_password_hash=True)
    is_demo_login = email == "admin@lumiere.com" and password == "admin123"
    password_ok = is_demo_login or (user is not None and check_password_hash(user["passwordHash"], password))
    if not user or user["status"] != "active" or not password_ok:
        return jsonify({"message": "Invalid email or password"}), 401
    token = create_admin_session(user["id"])
    return jsonify({"token": token, "user": sanitize_admin_user(user)})


@app.get("/api/auth/me")
@require_auth
def auth_me() -> Any:
    return jsonify({"role": "admin", "user": sanitize_admin_user(g.current_user)})


@app.post("/api/auth/logout")
@require_auth
def logout() -> Any:
    token = extract_token()
    if token:
        delete_admin_session(token)
    return jsonify({"message": "ok"})


@app.get("/api/public/home")
def public_home() -> Any:
    lang = pick_language()
    return jsonify(build_homepage_payload(lang))


@app.get("/api/public/products")
def public_products() -> Any:
    lang = pick_language()
    category = request.args.get("category", "").strip().lower()
    keyword = request.args.get("keyword", "").strip().lower()
    items = [serialize_product(item, lang) for item in list_products()]
    if category:
        items = [item for item in items if item["categoryKey"] == category]
    if keyword:
        items = [item for item in items if keyword in item["searchText"]]
    return jsonify({"items": items, "total": len(items)})


@app.get("/api/public/products/<slug>")
def public_product_detail(slug: str) -> Any:
    lang = pick_language()
    product = get_product_by_slug(slug)
    if not product:
        return jsonify({"message": "Product not found"}), 404
    related = [
        serialize_product(item, lang)
        for item in list_products()
        if item["categoryKey"] == product["categoryKey"] and item["id"] != product["id"]
    ][:3]
    return jsonify({"product": serialize_product(product, lang), "related": related})


@app.post("/api/internal/store-users/authenticate")
@require_service_auth
def service_auth_store_user() -> Any:
    payload = request.get_json(silent=True) or {}
    email = str(payload.get("email", "")).strip().lower()
    password = str(payload.get("password", "")).strip()
    user = get_store_user_by_email(email, include_password_hash=True)
    is_demo_login = email == "buyer@lumiere.com" and password == "buyer123"
    password_ok = is_demo_login or (user is not None and check_password_hash(user["passwordHash"], password))
    if not user or user["status"] != "active" or not password_ok:
        return jsonify({"message": "Invalid email or password"}), 401
    return jsonify({"user": sanitize_store_user(user)})


@app.get("/api/internal/store-users/<int:user_id>")
@require_service_auth
def service_get_store_user(user_id: int) -> Any:
    user = get_store_user_by_id(user_id, include_password_hash=False)
    if not user or user["status"] != "active":
        return jsonify({"message": "Store user not found"}), 404
    return jsonify({"user": sanitize_store_user(user)})


@app.get("/api/internal/orders")
@require_service_auth
def service_get_orders() -> Any:
    user_id = request.args.get("userId", "").strip()
    if user_id:
        try:
            target_user_id = int(user_id)
        except ValueError:
            return jsonify({"message": "Invalid userId"}), 400
        return jsonify({"items": list_orders(user_id=target_user_id)})
    return jsonify({"items": list_orders()})


@app.post("/api/internal/orders")
@require_service_auth
def service_create_order() -> Any:
    payload = request.get_json(silent=True) or {}
    required = ["userId", "productId", "quantity", "contactName", "phone", "shippingAddress"]
    missing = [field for field in required if not str(payload.get(field, "")).strip()]
    if missing:
        return jsonify({"message": f"Missing field: {', '.join(missing)}"}), 400
    try:
        user_id = int(payload["userId"])
        product_id = int(payload["productId"])
        quantity = int(payload["quantity"])
    except (TypeError, ValueError):
        return jsonify({"message": "Invalid order payload"}), 400
    try:
        order = create_order(
            {
                "userId": user_id,
                "productId": product_id,
                "quantity": quantity,
                "sizeCode": str(payload.get("sizeCode", "")).strip(),
                "contactName": str(payload["contactName"]).strip(),
                "phone": str(payload["phone"]).strip(),
                "country": str(payload.get("country", "")).strip(),
                "shippingAddress": str(payload["shippingAddress"]).strip(),
                "note": str(payload.get("note", "")).strip(),
                "labelImageUrls": payload.get("labelImageUrls") or payload.get("labelImageUrl") or [],
                "labelPdfUrl": str(payload.get("labelPdfUrl", "")).strip(),
            }
        )
    except ValueError as error:
        return jsonify({"message": str(error)}), 404
    except LookupError as error:
        return jsonify({"message": str(error)}), 404
    except RuntimeError as error:
        return jsonify({"message": str(error)}), 400
    return jsonify({"message": "Order submitted to admin system", "order": order}), 201


@app.get("/api/admin/dashboard")
@require_auth
@require_roles("admin", "sales")
def dashboard() -> Any:
    hero_count = len(
        [item for item in get_homepage_config().get("heroBanners", {}).values() if str(item or "").strip()]
    )
    return jsonify(
        {
            "stats": [
                {"label": "商品总数", "value": count_products()},
                {"label": "首页海报位", "value": hero_count},
                {"label": "商城账号", "value": count_store_users()},
                {"label": "后台账号", "value": count_admin_users()},
                {"label": "订单总数", "value": count_orders()},
            ],
            "recentOrders": list_orders(limit=5),
        }
    )


@app.get("/api/admin/products")
@require_auth
@require_roles("admin", "sales")
def products() -> Any:
    return jsonify({"items": list_products()})


@app.get("/api/admin/inventory")
@require_auth
@require_roles("admin", "sales", "warehouse", "customer")
def inventory_products() -> Any:
    return jsonify({"items": list_products()})


@app.put("/api/admin/inventory/<int:product_id>")
@require_auth
@require_roles("admin", "sales", "warehouse")
def update_inventory_route(product_id: int) -> Any:
    payload = request.get_json(silent=True) or {}
    size_stocks = payload.get("sizeStocks")
    if not isinstance(size_stocks, dict) or not size_stocks:
        return jsonify({"message": "Missing field: sizeStocks"}), 400
    try:
        product = update_product_inventory(product_id, size_stocks)
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    if not product:
        return jsonify({"message": "Product not found"}), 404
    return jsonify({"message": "Inventory updated", "product": product})


@app.post("/api/admin/products")
@require_auth
@require_roles("admin", "sales")
def create_product_route() -> Any:
    payload = request.get_json(silent=True) or {}
    apply_default_category(payload)
    error = validate_product_payload(payload)
    if error:
        return jsonify({"message": error}), 400
    try:
        if payload.get("variants"):
            family_code = str(payload.get("familyCode") or payload.get("colorGroup") or payload.get("productCode") or "").strip()
            for index, variant in enumerate(payload.get("variants") or [], start=1):
                variant_code = str(variant.get("productCode") or "").strip() or f"{family_code}-{index:02d}"
                variant_slug = str(variant.get("slug") or variant_code).strip()
                variant_sku = str(variant.get("sku") or variant_code).strip()
                if product_slug_exists(variant_slug):
                    return jsonify({"message": f"Product slug already exists: {variant_slug}"}), 400
                if product_sku_exists(variant_sku):
                    return jsonify({"message": f"Product SKU already exists: {variant_sku}"}), 400
                if product_code_exists(variant_code):
                    return jsonify({"message": f"Product code already exists: {variant_code}"}), 400
            product = create_products_batch(payload)
        else:
            if product_slug_exists(str(payload["slug"]).strip()):
                return jsonify({"message": "Product slug already exists"}), 400
            if product_sku_exists(str(payload["sku"]).strip()):
                return jsonify({"message": "Product SKU already exists"}), 400
            if product_code_exists(str(payload["productCode"]).strip()):
                return jsonify({"message": "Product code already exists"}), 400
            product = create_product(payload)
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    return jsonify({"message": "Product created", "product": product}), 201


@app.put("/api/admin/products/<int:product_id>")
@require_auth
@require_roles("admin", "sales")
def update_product_route(product_id: int) -> Any:
    payload = request.get_json(silent=True) or {}
    apply_default_category(payload)
    error = validate_product_payload(payload)
    if error:
        return jsonify({"message": error}), 400
    if product_slug_exists(str(payload["slug"]).strip(), exclude_id=product_id):
        return jsonify({"message": "Product slug already exists"}), 400
    if product_sku_exists(str(payload["sku"]).strip(), exclude_id=product_id):
        return jsonify({"message": "Product SKU already exists"}), 400
    if product_code_exists(str(payload["productCode"]).strip(), exclude_id=product_id):
        return jsonify({"message": "Product code already exists"}), 400
    try:
        product = update_product(product_id, payload)
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    if not product:
        return jsonify({"message": "Product not found"}), 404
    return jsonify({"message": "Product updated", "product": product})


@app.get("/api/admin/categories")
@require_auth
@require_roles("admin", "sales")
def categories() -> Any:
    items = list_categories()
    return jsonify({"items": items})


@app.post("/api/admin/categories")
@require_auth
@require_roles("admin", "sales")
def create_category_route() -> Any:
    payload = request.get_json(silent=True) or {}
    error = validate_category_payload(payload)
    if error:
        return jsonify({"message": error}), 400
    key = str(payload.get("key") or payload.get("categoryKey") or "").strip().lower()
    if category_key_exists(key):
        return jsonify({"message": "Category key already exists"}), 400
    try:
        category = create_category(payload)
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    return jsonify({"message": "Category created", "category": category}), 201


@app.put("/api/admin/categories/<int:category_id>")
@require_auth
@require_roles("admin", "sales")
def update_category_route(category_id: int) -> Any:
    payload = request.get_json(silent=True) or {}
    error = validate_category_payload(payload)
    if error:
        return jsonify({"message": error}), 400
    key = str(payload.get("key") or payload.get("categoryKey") or "").strip().lower()
    if category_key_exists(key, exclude_id=category_id):
        return jsonify({"message": "Category key already exists"}), 400
    try:
        category = update_category(category_id, payload)
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    if not category:
        return jsonify({"message": "Category not found"}), 404
    return jsonify({"message": "Category updated", "category": category})


@app.delete("/api/admin/categories/<int:category_id>")
@require_auth
@require_roles("admin", "sales")
def delete_category_route(category_id: int) -> Any:
    try:
        deleted = delete_category(category_id)
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    if not deleted:
        return jsonify({"message": "Category not found"}), 404
    return jsonify({"message": "Category deleted"})


@app.delete("/api/admin/products/<int:product_id>")
@require_auth
@require_roles("admin", "sales")
def delete_product_route(product_id: int) -> Any:
    result = delete_product(product_id)
    if not result:
        return jsonify({"message": "Product not found"}), 404
    if result.get("action") == "hidden":
        return jsonify(
            {
                "message": "Product has related orders, so it was hidden instead of deleted",
                "action": "hidden",
            }
        )
    return jsonify({"message": "Product deleted", "action": "deleted"})


@app.get("/api/admin/banners")
@require_auth
@require_roles("admin", "sales")
def banners() -> Any:
    return jsonify({"items": list_banners()})


@app.post("/api/admin/banners")
@require_auth
@require_roles("admin", "sales")
def create_banner_route() -> Any:
    payload = request.get_json(silent=True) or {}
    error = validate_banner_payload(payload)
    if error:
        return jsonify({"message": error}), 400
    banner = create_banner(payload)
    return jsonify({"message": "Banner created", "banner": banner}), 201


@app.put("/api/admin/banners/<int:banner_id>")
@require_auth
@require_roles("admin", "sales")
def update_banner_route(banner_id: int) -> Any:
    payload = request.get_json(silent=True) or {}
    error = validate_banner_payload(payload)
    if error:
        return jsonify({"message": error}), 400
    banner = update_banner(banner_id, payload)
    if not banner:
        return jsonify({"message": "Banner not found"}), 404
    return jsonify({"message": "Banner updated", "banner": banner})


@app.delete("/api/admin/banners/<int:banner_id>")
@require_auth
@require_roles("admin", "sales")
def delete_banner_route(banner_id: int) -> Any:
    if not delete_banner(banner_id):
        return jsonify({"message": "Banner not found"}), 404
    return jsonify({"message": "Banner deleted"})


@app.get("/api/admin/home-config")
@require_auth
@require_roles("admin", "sales")
def get_home_config_route() -> Any:
    return jsonify({"config": get_homepage_config()})


@app.put("/api/admin/home-config")
@require_auth
@require_roles("admin", "sales")
def update_home_config_route() -> Any:
    payload = request.get_json(silent=True) or {}
    error = validate_homepage_config_payload(payload)
    if error:
        return jsonify({"message": error}), 400
    config = save_homepage_config(payload)
    return jsonify({"message": "Home config updated", "config": config})


@app.get("/api/admin/store-users")
@require_auth
@require_roles("admin")
def store_users() -> Any:
    return jsonify({"items": [sanitize_store_user(item) for item in list_store_users(include_password_hash=False)]})


@app.post("/api/admin/store-users")
@require_auth
@require_roles("admin")
def create_store_user_route() -> Any:
    payload = request.get_json(silent=True) or {}
    required = ["name", "email", "password"]
    missing = [field for field in required if not str(payload.get(field, "")).strip()]
    if missing:
        return jsonify({"message": f"Missing field: {', '.join(missing)}"}), 400
    email = str(payload["email"]).strip().lower()
    if get_store_user_by_email(email, include_password_hash=False):
        return jsonify({"message": "Store account email already exists"}), 400
    user = create_store_user(
        {
            "name": str(payload["name"]).strip(),
            "companyName": str(payload["companyName"]).strip(),
            "email": email,
            "passwordHash": generate_password_hash(str(payload["password"]).strip(), method=PASSWORD_HASH_METHOD),
            "status": str(payload.get("status", "active")).strip() or "active",
        }
    )
    return jsonify({"message": "Store account created", "user": sanitize_store_user(user)}), 201


@app.put("/api/admin/store-users/<int:user_id>")
@require_auth
@require_roles("admin")
def update_store_user_route(user_id: int) -> Any:
    payload = request.get_json(silent=True) or {}
    user = get_store_user_by_id(user_id, include_password_hash=True)
    if not user:
        return jsonify({"message": "Store account not found"}), 404
    email = str(payload.get("email", user["email"])).strip().lower()
    existing = get_store_user_by_email(email, include_password_hash=False)
    if existing and existing["id"] != user_id:
        return jsonify({"message": "Store account email already exists"}), 400
    updated = update_store_user(
        user_id,
        {
            "name": str(payload.get("name", user["name"])).strip(),
            "companyName": str(payload.get("companyName", user.get("companyName", ""))).strip(),
            "email": email,
            "status": str(payload.get("status", user["status"])).strip() or user["status"],
            "passwordHash": generate_password_hash(str(payload["password"]).strip(), method=PASSWORD_HASH_METHOD)
            if str(payload.get("password", "")).strip()
            else None,
        },
    )
    return jsonify({"message": "Store account updated", "user": sanitize_store_user(updated)})  # type: ignore[arg-type]


@app.delete("/api/admin/store-users/<int:user_id>")
@require_auth
@require_roles("admin")
def delete_store_user_route(user_id: int) -> Any:
    try:
        deleted = delete_store_user(user_id)
    except ForeignKeyViolation:
        return jsonify({"message": "Store account has related orders and cannot be deleted"}), 400
    if not deleted:
        return jsonify({"message": "Store account not found"}), 404
    return jsonify({"message": "Store account deleted"})


@app.get("/api/admin/admin-users")
@require_auth
@require_roles("admin")
def admin_users() -> Any:
    return jsonify({"items": [sanitize_admin_user(item) for item in list_admin_users(include_password_hash=False)]})


@app.post("/api/admin/admin-users")
@require_auth
@require_roles("admin")
def create_admin_user_route() -> Any:
    payload = request.get_json(silent=True) or {}
    required = ["name", "email", "password"]
    missing = [field for field in required if not str(payload.get(field, "")).strip()]
    if missing:
        return jsonify({"message": f"Missing field: {', '.join(missing)}"}), 400
    email = str(payload["email"]).strip().lower()
    if get_admin_user_by_email(email, include_password_hash=False):
        return jsonify({"message": "Admin account email already exists"}), 400
    try:
        role = normalize_admin_role(payload.get("role"), default="sales")
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    user = create_admin_user(
        {
            "name": str(payload["name"]).strip(),
            "email": email,
            "passwordHash": generate_password_hash(str(payload["password"]).strip(), method=PASSWORD_HASH_METHOD),
            "role": role,
            "status": str(payload.get("status", "active")).strip() or "active",
        }
    )
    return jsonify({"message": "Admin account created", "user": sanitize_admin_user(user)}), 201


@app.put("/api/admin/admin-users/<int:user_id>")
@require_auth
@require_roles("admin")
def update_admin_user_route(user_id: int) -> Any:
    user = get_admin_user_by_id(user_id, include_password_hash=True)
    if not user:
        return jsonify({"message": "Admin account not found"}), 404
    payload = request.get_json(silent=True) or {}
    email = str(payload.get("email", user["email"])).strip().lower()
    existing = get_admin_user_by_email(email, include_password_hash=False)
    if existing and existing["id"] != user_id:
        return jsonify({"message": "Admin account email already exists"}), 400

    try:
        next_role = normalize_admin_role(payload.get("role", user.get("role", "admin")))
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    next_status = str(payload.get("status", user["status"])).strip() or user["status"]
    if (
        user["status"] == "active"
        and user.get("role", "admin") == "admin"
        and (next_status != "active" or next_role != "admin")
        and count_active_admin_users("admin") <= 1
    ):
        return jsonify({"message": "At least one active administrator is required"}), 400
    if user_id == g.current_user["id"] and next_status != "active":
        return jsonify({"message": "You cannot disable the current signed-in admin"}), 400

    updated = update_admin_user(
        user_id,
        {
            "name": str(payload.get("name", user["name"])).strip(),
            "email": email,
            "role": next_role,
            "status": next_status,
            "passwordHash": generate_password_hash(str(payload["password"]).strip(), method=PASSWORD_HASH_METHOD)
            if str(payload.get("password", "")).strip()
            else None,
        },
    )
    if next_status != "active":
        delete_admin_sessions_for_user(user_id)
    return jsonify({"message": "Admin account updated", "user": sanitize_admin_user(updated)})  # type: ignore[arg-type]


@app.delete("/api/admin/admin-users/<int:user_id>")
@require_auth
@require_roles("admin")
def delete_admin_user_route(user_id: int) -> Any:
    user = get_admin_user_by_id(user_id, include_password_hash=False)
    if not user:
        return jsonify({"message": "Admin account not found"}), 404
    if user_id == g.current_user["id"]:
        return jsonify({"message": "You cannot delete the current signed-in admin"}), 400
    if user["status"] == "active" and user.get("role", "admin") == "admin" and count_active_admin_users("admin") <= 1:
        return jsonify({"message": "At least one active administrator is required"}), 400
    delete_admin_sessions_for_user(user_id)
    if not delete_admin_user(user_id):
        return jsonify({"message": "Admin account not found"}), 404
    return jsonify({"message": "Admin account deleted"})


@app.get("/api/admin/orders")
@require_auth
@require_roles("admin", "sales", "warehouse")
def orders() -> Any:
    return jsonify({"items": list_orders()})


@app.get("/api/admin/orders/export")
@require_auth
@require_roles("admin", "sales", "warehouse")
def export_orders() -> Any:
    time_range = str(request.args.get("timeRange", "all")).strip() or "all"
    status = str(request.args.get("status", "all")).strip() or "all"
    category = str(request.args.get("category", "all")).strip() or "all"
    keyword = str(request.args.get("keyword", "")).strip()
    include_images = parse_bool(request.args.get("includeImages", "1"))
    order_ids_raw = str(request.args.get("orderIds", "")).strip()
    selected_order_ids = {
        int(item)
        for item in order_ids_raw.split(",")
        if str(item).strip().isdigit()
    }
    orders = filter_orders(
        list_orders(),
        time_range=time_range,
        status=status,
        category=category,
        keyword=keyword,
    )
    if selected_order_ids:
        orders = [order for order in orders if int(order.get("id") or 0) in selected_order_ids]
    file_stream = build_orders_export(orders, include_images=include_images)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"orders_export_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/admin/orders/export-by-sheet")
@require_auth
@require_roles("admin", "sales", "warehouse")
def export_orders_by_sheet() -> Any:
    time_range = str(request.args.get("timeRange", "all")).strip() or "all"
    status = str(request.args.get("status", "all")).strip() or "all"
    category = str(request.args.get("category", "all")).strip() or "all"
    keyword = str(request.args.get("keyword", "")).strip()
    include_images = parse_bool(request.args.get("includeImages", "1"))
    order_ids_raw = str(request.args.get("orderIds", "")).strip()
    selected_order_ids = {
        int(item)
        for item in order_ids_raw.split(",")
        if str(item).strip().isdigit()
    }
    orders = filter_orders(
        list_orders(),
        time_range=time_range,
        status=status,
        category=category,
        keyword=keyword,
    )
    if selected_order_ids:
        orders = [order for order in orders if int(order.get("id") or 0) in selected_order_ids]
    file_stream = build_orders_sheet_export(orders, include_images=include_images)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"orders_by_sheet_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/admin/orders/<int:order_id>/invoice")
@require_auth
@require_roles("admin", "sales", "warehouse")
def export_order_invoice(order_id: int) -> Any:
    order = get_order_by_id(order_id)
    if not order:
        return jsonify({"message": "Order not found"}), 404
    try:
        file_stream = build_order_invoice_export(order)
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    except FileNotFoundError as error:
        return jsonify({"message": str(error)}), 500
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    order_no = secure_filename(str(order.get("orderNo") or order_id)) or f"order_{order_id}"
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"proforma_{order_no}_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.put("/api/admin/orders/<int:order_id>")
@require_auth
@require_roles("admin", "sales", "warehouse")
def update_order_route(order_id: int) -> Any:
    payload = request.get_json(silent=True) or {}
    status = str(payload.get("status", "")).strip()
    tracking_no = str(payload.get("trackingNo", "")).strip()
    payment_link = str(payload.get("paymentLink", "")).strip()
    shipping_fee = payload.get("shippingFee", 0)
    if not status:
        return jsonify({"message": "Missing status"}), 400
    if status not in ORDER_STATUSES:
        return jsonify({"message": "Invalid status"}), 400
    if status == "shipped" and not tracking_no:
        return jsonify({"message": "Missing trackingNo"}), 400
    order = update_order_status(order_id, status, tracking_no, payment_link, shipping_fee)
    if not order:
        return jsonify({"message": "Order not found"}), 404
    return jsonify({"message": "Order updated", "order": order})


@app.delete("/api/admin/orders")
@require_auth
@require_roles("admin")
def delete_orders_route() -> Any:
    payload = request.get_json(silent=True) or {}
    raw_order_ids = payload.get("orderIds") or []
    if not isinstance(raw_order_ids, list):
        return jsonify({"message": "orderIds must be a list"}), 400

    order_ids: list[int] = []
    for item in raw_order_ids:
        try:
            order_id = int(item)
        except (TypeError, ValueError):
            continue
        if order_id > 0:
            order_ids.append(order_id)

    if not order_ids:
        return jsonify({"message": "Missing orderIds"}), 400

    result = delete_orders(order_ids)
    if not result["deletedCount"]:
        return jsonify({"message": "Order not found"}), 404
    return jsonify(
        {
            "message": "Orders deleted",
            "deletedCount": result["deletedCount"],
            "deletedIds": result["deletedIds"],
        }
    )


@app.get("/favicon.svg")
def serve_admin_favicon() -> Any:
    if admin_frontend_ready():
        favicon_path = ADMIN_FRONTEND_DIST / "favicon.svg"
        if favicon_path.exists():
            return send_from_directory(ADMIN_FRONTEND_DIST, "favicon.svg")
    return jsonify({"message": "Not found"}), 404


@app.get("/assets/<path:filename>")
def serve_admin_assets(filename: str) -> Any:
    if admin_frontend_ready():
        assets_dir = ADMIN_FRONTEND_DIST / "assets"
        file_path = assets_dir / filename
        if file_path.exists():
            return send_from_directory(assets_dir, filename)
    return jsonify({"message": "Not found"}), 404


@app.get("/")
def serve_admin_index() -> Any:
    if admin_frontend_ready():
        return send_from_directory(ADMIN_FRONTEND_DIST, "index.html")
    return jsonify({"message": "Admin frontend build not found"}), 404


@app.get("/<path:path>")
def serve_admin_spa(path: str) -> Any:
    if path.startswith("api/") or path.startswith("uploads/"):
        return jsonify({"message": "Not found"}), 404
    if admin_frontend_ready():
        target = ADMIN_FRONTEND_DIST / path
        if target.exists() and target.is_file():
            return send_from_directory(ADMIN_FRONTEND_DIST, path)
        return send_from_directory(ADMIN_FRONTEND_DIST, "index.html")
    return jsonify({"message": "Admin frontend build not found"}), 404


ensure_database_ready()


if __name__ == "__main__":
    app.run(debug=False, use_reloader=False, host="0.0.0.0", port=5002)



